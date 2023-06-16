# 使用xlsx檔作為儲存
import datetime
import re
import openpyxl
from flask import Flask, request, abort
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import MessageEvent, TextMessage

# 建立flask物件
app = Flask(__name__)

# 導入linebot的TOKEN
Line_Bot_Api = LineBotApi("your_acces_token")
# 導入acces ID
handler = WebhookHandler("channel_acces")

# 要執行的指令
action = ""
# 建立一個日期變數
day = datetime.datetime.today().strftime

# 設置路由
@app.route("/callback", methods=['POST'])
# 接收linebot的Webhook事件
def callback():
    # 取得line平台的簽名資訊
    signture = request.headers['X-Line-Signature']
    # 獲取請求的內容(以文字型態回傳)
    body = request.get_data(as_text=True)
    try:
        # 接收Line Bot事件
        handler.handle(body, signture)
    # 驗證失敗的話
    except InvalidSignatureError:
        # 回傳一個400的狀態碼
        abort(400)
    return 'OK'

# 建立一個裝飾器，用於將下面的函數註冊為Line Bot SDK中處理特定事件的處理器
@handler.add(MessageEvent, message=TextMessage)
def hamlle_message(event):
    # 將action設定為全域變數
    global action
    # 取得傳入的文字
    msg = event.message.text

    # 檢查是不是指令
    regex = re.compile("\[(.+)\]")
    match = regex.match(msg)

    # 判斷指令，不是指令就忽略
    if match:
        # 如果是指令的話，設定下一步要執行的指令
        action = match.group(1)
        # 取得回傳訊息
        reply_msg = getActionReplyMsg()

        # 回傳訊息
        Line_Bot_Api.reply_message(
            event.reply_token,
            TextMessage(text=reply_msg)
        )
        # 結束本次執行
        return
    else:
        # 執行收入的選項
        if action == '收入':
            
            # 設定回覆訊息
            reply_msg = (
                day('%m%d') + '收入︰' + msg
                )
            # 設定機器人並載入token,回覆訊息
            Line_Bot_Api.reply_message(
                event.reply_token,
                TextMessage(text=reply_msg)
            )
            # 執行存檔的函數
            save_data(msg)
            # 脫離指令迴圈
            action = ""

        # 執行支出的選項
        elif action == '支出':
            reply_msg = (
                day('%m%d') + '支出︰' + msg
            )
            Line_Bot_Api.reply_message(
                event.reply_token,
                TextMessage(text=reply_msg)
            )
            save_data(msg)
            action = ""

        # 執行月收入的選項
        elif action == '月收入':
            if int(msg) not in [
                1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12
                ]:
                reply_msg = (
                    '輸入錯誤'
                )
            else:
                if not check_data(msg):
                    if load_data(msg) != False:
                        month_msg = day('%m') + '月收入︰' + load_data(msg)
                    else:
                        month_msg = f'本月尚未有收入紀錄'
                else:
                    month_msg = '查無該月資料'

                reply_msg = (
                    month_msg
                )
                    
            Line_Bot_Api.reply_message(
                event.reply_token,
                TextMessage(text=reply_msg)
            )
            action = ""

        # 執行月支出的選項
        elif action == '月支出':
            if int(msg) not in [
                1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12
                ]:
                reply_msg = (
                    '輸入錯誤'
                )
            else:
                if not check_data(msg):
                    if load_data(msg) != False:
                        month_msg = day('%m') + '月支出︰' + load_data(msg)
                    else:
                        month_msg = f'本月尚未有支出紀錄'
                else:
                    month_msg = '查無該月資料'

                reply_msg = (
                    month_msg
                )
                    
            Line_Bot_Api.reply_message(
                event.reply_token,
                TextMessage(text=reply_msg)
            )
            action = ""
        
# 確認收到的指令內容
def getActionReplyMsg():
    global action
    if action == '收入':
        return day('%m%d') + '-收入︰'
    elif action == '支出':
        return day('%m%d') + '-支出︰'
    elif action == '月收入':
        return '請輸入要查詢的月份︰'
    elif action == '月支出':
        return '請輸入要查詢的月份︰'
    elif action == '總收入':
        return day('%Y') + '-總收入︰'
    elif action == '總支出':
        return day('%Y') + '-總支出︰'
    else:
        return '無效指令'

# 建立一個存檔的函數
def save_data(money):
    today = day('%d')
    # 檢查檔案是否存在
    try:
        # 開始Excel表
        wb = openpyxl.load_workbook('data/money.xlsx', data_only=True)
    # 發現不存在
    except FileNotFoundError:
        # 建立一個工作表
        wb = openpyxl.Workbook()
        # 建立一個活頁簿
        s1 = wb.create_sheet(f'{day("%m")}月', 0)
        # 預設第一行第一列為日期
        s1.cell(1, 1).value = '日期'
        # 預設第一行第二列為收入
        s1.cell(1, 2).value = '收入'
        # 預設第一行第三列為支出
        s1.cell(1, 3).value = '支出'
        # 預設第一行第四列為本日收支
        s1.cell(1, 4).value = '收支'
    # 如果存在則繼續執行
    else:
        # 開啟當月的活頁簿
        s1 = wb[f'{day("%m")}月']

    # 將row設為列的最大上限+1
    row = s1.max_row + 1
    # 建立一個判斷指令是收入還是支出
    if action == '收入':
        cul = 2
    elif action == '支出':
        cul = 3

    # 使用迴圈尋找是否有當天的值
    for i in range(1, row+1):
        # 確認當天是否有輸入
        if s1.cell(i, 1).value == today:
            # 當天有輸入並且有金額的時候
            if s1.cell(i, cul).value != None:
                # 取出原本的金額
                get_value = s1.cell(i, cul).value
                # 將原本的金額與新的金額相加
                s1.cell(i, cul).value = int(get_value) + int(money)
            # 如果沒有金額的話
            else:
                # 將金額直接建立
                s1.cell(i, cul).value = int(money)

            break
        # 當i=最大值時
        elif i == row:
            # 新增一個第i行第一列為日期
            s1.cell(row, 1).value = today
            # 新增一個第i行第cul列為收入或支出的金額
            s1.cell(row, cul).value = int(money)

    # 存檔
    wb.save('data/money.xlsx')

# 建立一個讀取的函數
def load_data(month):
    # 檢查輸入是否為兩位數，若不是的話則前面補0
    if len(month) <= 1:
        month = '0' + str(month)
    # 載入檔案
    wb = openpyxl.load_workbook('data/money.xlsx', data_only=True)
    # 載入活頁簿
    s1 = wb[f'{month}月']
    # 設定row為欄的最大數
    row = s1.max_row
    # 檢查目前指令為月收入或月支出
    if action == '月收入':
        cul = 2
    elif action == '月支出':
        cul = 3
    # 如果檔案內有資料的話
    if row > 2:
        # 將money1取得值
        money1 = int(s1.cell(2, cul).value)
        # 使用迴圈計算全部格數相加
        for i in range(3, row+1):
            # 將money2設為下一格的資料
            money2 = int(s1.cell(i, cul).value)
            # 將money1和money2相加
            money1 = money1 + money2
    # 如果資料只有一筆
    elif row == 2:
        # 將該筆資料直接回傳
        money1 = int(s1.cell(2, cul).value)
    # 如果沒有資料
    else:
        return False
        
    return str(money1)
        
# 建立一個檢查檔案的函數
def check_data(month):
    # 確認輸入是否為兩位數
    if len(month) <= 1:
        # 若不是兩位數則前面補0
        month = '0' + str(month)
    # 使用try-except檢查檔案是否存在
    try:
        wb = openpyxl.load_workbook('data/money.xlsx', data_only=True)
    # 如果找不到檔案回傳false
    except FileNotFoundError:
        return True
    # 如果找到檔案就繼續
    else:
        # 檢查活頁簿是否存在
        try:
            wb[f'{month}月']
        # 如果不存在回傳True
        except KeyError:
            return True
        # 如果存在回傳False
        else:
            return False

if __name__ == '__main__':
    app.run()
