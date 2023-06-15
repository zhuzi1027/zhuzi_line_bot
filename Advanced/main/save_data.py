import datetime
import openpyxl

today = datetime.datetime.today().strftime('%d')
month = datetime.datetime.today().strftime('%m')
class save():
    def day_data(money, action):
        try:
            wb = openpyxl.load_workbook('data/money1.xlsx', data_only=True)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            s1 = wb.create_sheet(f'{month}月')
            s1.cell(1, 1).value = '日期'
            s1.cell(1, 2).value = '收入'
            s1.cell(1, 3).value = '支出'
            s1.cell(1, 4).value = '收支'
        else:
            s1 = wb[f'{month}月']

        row = s1.max_row + 1
        if action == '收入':
            cul = 2
        else:
            cul = 3
        
        for i in range(1, row+1):
            if s1.cell(i, 1).value == today:
                if s1.cell(i, cul).value != None:
                    get_value = s1.cell(i, cul).value
                    s1.cell(i, cul).value = int(get_value) + int(money)
                else:
                    s1.cell(i, cul).value = int(money)
                break
            elif i == row:
                s1.cell(row, 1).value = today
                s1.cell(row, cul).value = int(money)
    
        wb.save('data/money1.xlsx')