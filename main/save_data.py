import datetime
import openpyxl

today = datetime.datetime.today().strftime('%d')
month = datetime.datetime.today().strftime('%m')
year = datetime.datetime.today().strftime('%Y')

class save():
    def day_data(money, action):
        if action == '收入':
            cul = 2
            right_money = int(money)
        else:
            cul = 3
            right_money = int(money) * (-1)

        try:
            wb = openpyxl.load_workbook(f'data/{year[2:]}.xlsx', data_only=True)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            s1 = wb.create_sheet(f'{month}月', 1)
            s1.cell(1, 1).value = '日期'
            s1.cell(1, 2).value = '收入'
            s1.cell(1, 3).value = '支出'
            s1.cell(1, 4).value = '收支'
        else:
            s1 = wb[f'{month}月']

        row_month = s1.max_row + 1
        
        for i in range(2, row_month+1):
            if s1.cell(i, 1).value == today:
                if s1.cell(i, cul).value != None:
                    get_value = s1.cell(i, cul).value
                    s1.cell(i, cul).value = int(get_value) + int(money)
                    s1.cell(i, 4).value = int(s1.cell(i, 4).value) + right_money
                else:
                    s1.cell(i, cul).value = int(money)
                    if s1.cell(i, 4).value == None:
                        s1.cell(i, 4).value = int(right_money)
                    else:
                        s1.cell(i, 4).value = int(s1.cell(i, 4).value) + right_money
                break
            elif i == row_month:
                s1.cell(row_month, 1).value = today
                s1.cell(row_month, cul).value = int(money)
                if s1.cell(i, 4).value == None:
                    s1.cell(i, 4).value = int(right_money)
                else:
                    s1.cell(i, 4).value = int(s1.cell(i, 4).value) + right_money

        try:
            wb['年收支']
        except KeyError:
            s2 = wb.create_sheet('年收支', 0)
            s2.cell(1, 1).value = '月份'
            s2.cell(1, 2).value = '收入'
            s2.cell(1, 3).value = '支出'
            s2.cell(1, 4).value = '收支'
        else:
            s2 = wb['年收支']
        
        row_year = s2.max_row+1

        for j in range(2, row_year+1):
            if s2.cell(j, 1).value == month:
                if s2.cell(j, cul).value != None:
                    get_value = s2.cell(j, cul).value
                    s2.cell(j, cul).value = int(get_value) + int(money)
                    s2.cell(j, 4).value = int(s2.cell(j, 4).value) + right_money
                else:
                    s2.cell(j, cul).value = int(money)
                    if s2.cell(i, 4).value == None:
                        s2.cell(i, 4).value = int(right_money)
                    else:
                        s2.cell(i, 4).value = int(s2.cell(i, 4).value) + right_money
                break
            elif j == row_year:
                s2.cell(j, 1).value = month
                s2.cell(j, cul).value = int(money)
                if s2.cell(i, 4).value == None:
                    s2.cell(i, 4).value = int(right_money)
                else:
                    s2.cell(i, 4).value = int(s2.cell(i, 4).value) + right_money

        wb.save(f'data/{year[2:]}.xlsx')