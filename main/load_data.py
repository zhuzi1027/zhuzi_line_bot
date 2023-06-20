import datetime
import openpyxl

today = datetime.datetime.today().strftime('%d')
year = datetime.datetime.today().strftime('%Y')

class load():
    def check_data(month):
        if len(month) < 2:
            month = '0' + str(month)
        try:
            wb = openpyxl.load_workbook(f'data/{year[2:]}.xlsx', data_only=True)
            wb['年收支']
        except FileNotFoundError:
            return False
        except KeyError:
            return False
        else:
            return True
            
    def load_data(month, action):
        if len(month) < 2:
            month = '0' + str(month)
        wb = openpyxl.load_workbook(f'data/{year[2:]}.xlsx', data_only=True)
        s1 = wb['年收支']
        row = s1.max_row
        if action == '月收入':
            cul = 2
        elif action == '月支出':
            cul = 3
        
        for i in range(2, row+1):
            if s1.cell(i, 1).value == month:
                if s1.cell(i, cul).value != None:
                    money1 = s1.cell(i, cul).value
                else:
                    return False
            else:
                return False
        return str(money1)
    
    def load_year_data(year_msg, action):
        if action == '年收入':
            cul = 2
        else:
            cul = 3
        wb = openpyxl.load_workbook(f'data/{year_msg}.xlsx', data_only=True)
        s1 = wb['年收支']
        tota_value = 0
        tota_money = 0
        row = s1.max_row
        for i in range(2, row+1):
            value1 = s1.cell(i, cul).value
            if value1 != None:
                tota_value = tota_value + int(value1)
            else:
                pass
            money1 = s1.cell(i, 4).value
            if money1 != None:
                tota_money = tota_money + int(money1)
            else:
                pass

        return [str(tota_value), str(tota_money)]
        